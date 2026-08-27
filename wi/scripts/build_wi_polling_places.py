#!/usr/bin/env python3
"""
Build data/app/<county-slug>-polling-places.json for all 72 Wisconsin counties
— the STATEWIDE ward -> polling place pairing the ward card has never been
able to answer.

WHY THIS EXISTS. The ward card has shipped since phase 2 with the honest
answer "we cannot tell you where you vote": Wisconsin aggregates the pairing
only behind MyVote, whose every route sits behind a Cloudflare challenge this
project does not defeat (gap `ward-polling-places`). Two cities publish their
own pairing and ship from it — Milwaukee (build_mke_polling_places.py) and
Madison (build_madison_polling_places.py) — and the other 1,844
municipalities got a MyVote link.

The block was never a data floor, and the ask ledger proved it. Asked
directly, the Wisconsin Elections Commission sent the file: Jodi Vitcenda,
2026-08-27, help-desk ticket 123582, "Polling Place Listing (3).xlsx",
3,623 reporting units for the 3 November 2026 General Election.
`wi/data/source/wec/` holds it and the whole arrangement. A CHALLENGE ON A
WEBSITE IS NOT A REFUSAL BY THE AGENCY BEHIND IT — the same lesson Knox,
Johnson and Perry taught county-side, arriving here state-side.

IT IS PROVISIONAL AND THE CARD SAYS SO IN THE READER'S OWN SENTENCE. The
Commission's words: "I can't provide any direct links for the attachments at
this time, as it isn't published there for November yet, but that is where it
would be." So `_source.provisional` is true, per FILE and never per record,
and the module renders the word "provisional" in prose rather than as a badge.
Re-pull after 17 September 2026 (wi/WATCH.md) and clear it.

THE JOIN, AND WHY THE APP'S OWN FABRIC IS THE AUTHORITY. WEC keys a row by
(County, Muni, Reporting Unit) where a reporting unit is a ward LIST — "Ward
1", "Wards 1-3", "Wards 1-2,9-10". The app renders LTSB's ward fabric, so the
build walks LTSB's 7,161 wards and asks WEC for each; a WEC ward LTSB does
not carry can never be selected and is dropped, and an LTSB ward WEC does not
list gets no polling place and the card degrades to MyVote. Measured
2026-08-27: 7,131 of 7,161 paired (99.58%), the 30 misses enumerated below
and printed by every run.

FOUR TRAPS, ALL MEASURED, ALL ENCODED:

  1. A RANGE'S ENDPOINTS DECIDE HOW IT EXPANDS. Wisconsin has 26 lettered
     wards (Wauwatosa's 2A/2B, Green Bay's 11A, ...). Chippewa Falls files
     "Wards 1-2" and holds wards 1, 1A, 2 — that range is {1,2}, NOT {1,1A,2}.
     Town of Wolf River files "Wards 1-2A" and holds 1, 2A, 2B — that range IS
     {1,2A}. So: both endpoints plain -> expand NUMERICALLY over plain numbers;
     either endpoint lettered -> expand over the municipality's own ordered
     ward list. Read one way both witnesses agree; read either way alone, one
     of them silently mis-assigns a polling place.
  2. RANGES RUN BACKWARDS. "Wards 8-7" (Chippewa Falls), "Wards 2-1" (Town of
     Morrison). Sorted, not rejected — and counted in the report.
  3. NAMES DISAGREE FIVE WAYS, four of which normalization cannot reach.
     Punctuation and case are handled by `norm` (LTSB "St Croix" vs WEC
     "ST. CROIX"; LTSB "Land O'Lakes" vs WEC "LAND O-LAKES"). The other four
     are different NAMES and sit in MUNI_ALIASES, each with its counterpart.
  4. A CROSS-COUNTY MUNICIPALITY'S COUNTY SPLIT IS NOT AGREED. WEC files the
     City of Brodhead's ward 1 under Green County; LTSB files it under Rock.
     For the 58 municipalities WEC itself labels "MULTIPLE COUNTIES" — WEC
     stating they are one municipality — a county-key miss falls back to the
     municipality-wide key. It is scoped to those 58 on purpose: Town of Unity
     exists in BOTH Clark and Trempealeau and they are different towns, so an
     unscoped name fallback would put one town's electors in the other's
     polling place.

THE SHAPE IS FLAT AND PER COUNTY, for two reasons that are not aesthetic.
Per county, because the card already computes its county's slug for the gaps
panel's outlines, and a reader fetches ~9 KB instead of ~950 KB. Flat
(ward key -> place record, the shape both city files already use), because
validate_index.py floors a roster file by its top-level key count — so each
of the 72 files carries its own measured floor, which a {places, wards}
envelope would reduce to a floor of 2.

Usage:
    python3 wi/scripts/build_wi_polling_places.py
    python3 wi/scripts/build_wi_polling_places.py --check   # CI drift gate
    python3 wi/scripts/build_wi_polling_places.py --xlsx /path/to/newer.xlsx
"""

import argparse
import collections
import datetime
import json
import os
import re
import subprocess
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INSTANCE_ROOT = os.path.dirname(SCRIPT_DIR)
APP_DATA_DIR = os.path.join(INSTANCE_ROOT, "data", "app")
DEFAULT_XLSX = os.path.join(INSTANCE_ROOT, "data", "source", "wec",
                            "Polling Place Listing (3).xlsx")

LTSB_WARDS = ("https://services1.arcgis.com/FDsAtKBk8Hy4cAH0/arcgis/rest/"
              "services/WI_Municipal_Wards_Current/FeatureServer/0/query"
              "?where=1%3D1&outFields=WARDID,MCD_NAME,CTV,CNTY_NAME"
              "&returnGeometry=false&f=json&resultRecordCount=2000&resultOffset=")

# What the Commission sent, stated once. `provisional` is what the card says
# out loud; `electionDate` is what makes the card RETIRE itself rather than
# keep showing a past election's places (contract condition 5).
ELECTION = "2026 General Election"
ELECTION_DATE = "2026-11-03"
PUBLISHER = "Wisconsin Elections Commission"
SOURCE_URL = "https://elections.wi.gov/elections"
PROVISIONAL = True

EXPECT_COUNTIES = 72
EXPECT_LTSB_WARDS = 7161      # the July 2026 filing edition; see wi/WATCH.md
MIN_PAIR_RATE = 0.99          # measured 0.9958 — a drop past this is a broken join
WI_BBOX = {"min_lat": 42.4, "max_lat": 47.4, "min_lng": -93.0, "max_lng": -86.7}

# Trap 3: four municipalities WEC and LTSB spell with DIFFERENT WORDS, not
# different punctuation. Each maps a normalized WEC key onto its normalized
# LTSB counterpart; anything reachable by `norm` is deliberately NOT here.
MUNI_ALIASES = {
    ("CRAWFORD", "V", "MT STERLING"): ("CRAWFORD", "V", "MOUNT STERLING"),
    ("SAUK", "V", "LAVALLE"): ("SAUK", "V", "LA VALLE"),
    ("WALWORTH", "V", "FONTANA"): ("WALWORTH", "V", "FONTANAONGENEVA LAKE"),
    ("WAUPACA", "T", "SAINT LAWRENCE"): ("WAUPACA", "T", "ST LAWRENCE"),
}

CTV_OF = {"TOWN": "T", "CITY": "C", "VILLAGE": "V"}


def norm(s):
    """The one normalization, and the JS module in wi/index.html runs the SAME
    three rules on LTSB's live attributes to build its lookup key: uppercase,
    delete . ' and -, collapse whitespace. Change it here and change it there,
    or every ward in the state misses."""
    s = str(s).upper()
    for ch in (".", "'", "-"):
        s = s.replace(ch, "")
    return re.sub(r"\s+", " ", s).strip()


def slug_of(county_name):
    """"St Croix" -> "st-croix" — the same slug build_wi_county_outlines.py
    writes, so the card's one county slug addresses both files."""
    return re.sub(r"[^a-z0-9]+", "-", str(county_name).replace(".", "")
                  .replace("'", "").lower()).strip("-")


def ward_norm(raw):
    """"0011" -> "11", "011A" -> "11A". Zero padding is LTSB's; the letter is
    the ward's own identity and is NEVER dropped (parseInt would)."""
    m = re.fullmatch(r"0*(\d+)([A-Z]*)", str(raw).strip().upper())
    return (m.group(1) + m.group(2)) if m else str(raw).strip().upper()


def ward_sort(w):
    m = re.fullmatch(r"(\d+)([A-Z]*)", w)
    return (int(m.group(1)), m.group(2)) if m else (10 ** 6, w)


def fetch(url, tries=6, timeout=120):
    last = None
    for _ in range(tries):
        r = subprocess.run(
            ["curl", "-sSL", "--fail", "--max-time", str(timeout),
             "-H", "User-Agent: districtry/1.0 (+https://districtry.com/wi/)",
             url], capture_output=True, text=True)
        if r.returncode == 0:
            return r.stdout
        last = r.stderr.strip()
    raise SystemExit("fetch failed after %d tries: %s\n  %s" % (tries, url, last))


def load_ltsb():
    """LTSB's live statewide ward fabric — the same layer the app renders, so
    the pairing is gated against what a reader can actually select."""
    rows, offset = [], 0
    while True:
        d = json.loads(fetch(LTSB_WARDS + str(offset)))
        if "error" in d:
            raise SystemExit("LTSB ward layer error: %r" % d["error"])
        feats = d.get("features") or []
        rows += [f["attributes"] for f in feats]
        if not d.get("exceededTransferLimit") or not feats:
            break
        offset += len(feats)
    if len(rows) != EXPECT_LTSB_WARDS:
        print("  NOTE: LTSB now carries %d wards (was %d at build time) — a "
              "filing window has passed; re-read wi/WATCH.md's ward row"
              % (len(rows), EXPECT_LTSB_WARDS), file=sys.stderr)
    return rows


def read_wec(path):
    """The Commission's workbook, one row per reporting unit."""
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("openpyxl is required: pip install -c "
                         "wi/scripts/requirements.txt openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if wb.sheetnames != ["Table"]:
        raise SystemExit("unexpected sheets %r — the Commission's export shape "
                         "moved; re-measure before trusting the column map"
                         % (wb.sheetnames,))
    rows = list(wb["Table"].iter_rows(values_only=True))
    wb.close()
    header = list(rows[0])
    expect = ["County", "Muni", "Reporting Unit", "HINDI", "Polling Place Name",
              "Polling Place Address", "Voting Room Area", "Type Of Building",
              "Latitude", "Longitude", "Election Name", "Date Of Last Audit",
              "PPLID"]
    if header != expect:
        raise SystemExit("column header changed:\n  got      %r\n  expected %r"
                         % (header, expect))
    idx = {h: n for n, h in enumerate(header)}
    body = [r for r in rows[1:] if any(v is not None for v in r)]
    return idx, body


def parse_muni(raw):
    """"TOWN OF ADAMS - ADAMS COUNTY" -> ("T", "ADAMS"). The suffix after
    " - " is the county OR the literal "MULTIPLE COUNTIES", so it is never the
    key; the County column is. Case is inconsistent across rows ("Village of
    RIB MOUNTAIN" beside "VILLAGE OF ..."), hence the case-insensitive match."""
    base = str(raw).split(" - ")[0].strip()
    m = re.match(r"^(TOWN|CITY|VILLAGE)\s+OF\s+(.+)$", base, flags=re.I)
    if not m:
        raise SystemExit("unparseable municipality %r" % raw)
    return CTV_OF[m.group(1).upper()], norm(m.group(2))


def expand_reporting_unit(raw, ward_order):
    """Trap 1 and trap 2 live here. Returns (wards, notes)."""
    s = re.sub(r"^Wards?\s+", "", str(raw).strip(), flags=re.I)
    out, notes = [], []
    for part in [p.strip() for p in s.split(",")]:
        m = re.fullmatch(r"(\d+)([A-Z]?)\s*-\s*(\d+)([A-Z]?)", part, flags=re.I)
        if m:
            a, al, b, bl = m.group(1), m.group(2).upper(), m.group(3), m.group(4).upper()
            if not al and not bl:
                lo, hi = int(a), int(b)
                if lo > hi:
                    lo, hi = hi, lo
                    notes.append("reversed range %r" % part)
                out += [str(n) for n in range(lo, hi + 1)]
            else:
                lo_w, hi_w = ward_norm(a + al), ward_norm(b + bl)
                if not ward_order or lo_w not in ward_order or hi_w not in ward_order:
                    notes.append("lettered range %r has no anchor in the "
                                 "municipality's ward list" % part)
                    continue
                i_lo, i_hi = ward_order.index(lo_w), ward_order.index(hi_w)
                if i_lo > i_hi:
                    i_lo, i_hi = i_hi, i_lo
                    notes.append("reversed range %r" % part)
                out += ward_order[i_lo:i_hi + 1]
            continue
        m2 = re.fullmatch(r"(\d+[A-Z]?)", part, flags=re.I)
        if m2:
            out.append(ward_norm(m2.group(1)))
        else:
            notes.append("unreadable ward token %r" % part)
    return out, notes


def build(xlsx_path):
    ltsb = load_ltsb()
    county_raw = {}
    ltsb_wards = collections.defaultdict(set)
    for w in ltsb:
        c = norm(w["CNTY_NAME"])
        county_raw[c] = w["CNTY_NAME"]
        ltsb_wards[(c, w["CTV"], norm(w["MCD_NAME"]))].add(ward_norm(w["WARDID"]))
    if len(county_raw) != EXPECT_COUNTIES:
        raise SystemExit("LTSB names %d counties, expected %d"
                         % (len(county_raw), EXPECT_COUNTIES))
    # A normalization collision would silently merge two municipalities'
    # polling places inside one county — check it rather than assume it.
    seen = {}
    for (c, ctv, m) in ltsb_wards:
        if (c, ctv, m) in seen:
            raise SystemExit("normalized key %r is claimed twice" % ((c, ctv, m),))
        seen[(c, ctv, m)] = True
    ward_order = {k: sorted(v, key=ward_sort) for k, v in ltsb_wards.items()}

    idx, body = read_wec(xlsx_path)
    elections = {str(r[idx["Election Name"]]) for r in body}
    if elections != {ELECTION}:
        raise SystemExit("workbook covers %r, this build declares %r — reconcile "
                         "ELECTION/ELECTION_DATE before writing"
                         % (sorted(elections), ELECTION))

    places, by_county, by_muni = {}, collections.defaultdict(dict), collections.defaultdict(dict)
    notes = []
    conflicts = []
    for r in body:
        pid = str(r[idx["PPLID"]])
        lat, lng = float(r[idx["Latitude"]]), float(r[idx["Longitude"]])
        if not (WI_BBOX["min_lat"] <= lat <= WI_BBOX["max_lat"]
                and WI_BBOX["min_lng"] <= lng <= WI_BBOX["max_lng"]):
            raise SystemExit("polling place %s (%s) sits outside Wisconsin at "
                             "%.5f,%.5f" % (pid, r[idx["Polling Place Name"]], lat, lng))
        rec = {"name": re.sub(r"\s+", " ", str(r[idx["Polling Place Name"]]).strip()),
               "address": re.sub(r"\s+", " ", str(r[idx["Polling Place Address"]]).strip()),
               "lat": round(lat, 6), "lng": round(lng, 6)}
        if pid in places and places[pid] != rec:
            raise SystemExit("PPLID %s describes two different places — it is "
                             "not the place key this build assumes" % pid)
        places[pid] = rec

        county = norm(re.sub(r"(?i)\s+COUNTY$", "", str(r[idx["County"]])))
        ctv, mcd = parse_muni(r[idx["Muni"]])
        key = MUNI_ALIASES.get((county, ctv, mcd), (county, ctv, mcd))
        multi = "MULTIPLE COUNTIES" in str(r[idx["Muni"]]).upper()
        wards, n = expand_reporting_unit(r[idx["Reporting Unit"]], ward_order.get(key))
        for note in n:
            notes.append("%s %s: %s" % (r[idx["County"]], r[idx["Muni"]], note))
        for w in wards:
            if by_county[key].get(w) not in (None, pid):
                conflicts.append((key, w, r[idx["Reporting Unit"]]))
            by_county[key][w] = pid
            if multi:
                by_muni[(ctv, key[2])][w] = pid
    if conflicts:
        for c in conflicts[:10]:
            print("  %s ward %s claimed twice (%s)" % c, file=sys.stderr)
        raise SystemExit("%d ward(s) assigned two different polling places — "
                         "the reporting-unit expansion is wrong, not the data"
                         % len(conflicts))

    out = {slug_of(n): {} for n in county_raw.values()}
    paired = fallback = 0
    unpaired = collections.defaultdict(list)
    for w in ltsb:
        c, ctv, m = norm(w["CNTY_NAME"]), w["CTV"], norm(w["MCD_NAME"])
        wid = ward_norm(w["WARDID"])
        pid = by_county.get((c, ctv, m), {}).get(wid)
        if pid is None:
            pid = by_muni.get((ctv, m), {}).get(wid)   # trap 4
            if pid is not None:
                fallback += 1
        if pid is None:
            unpaired[(c, ctv, m)].append(wid)
            continue
        paired += 1
        out[slug_of(county_raw[c])]["%s|%s|%s" % (ctv, m, wid)] = places[pid]

    rate = paired / float(len(ltsb))
    if rate < MIN_PAIR_RATE:
        raise SystemExit("only %d of %d LTSB wards (%.2f%%) paired to a polling "
                         "place — floor is %.0f%%; the join broke"
                         % (paired, len(ltsb), rate * 100, MIN_PAIR_RATE * 100))
    empty = [s for s, d in out.items() if not d]
    if empty:
        raise SystemExit("county file(s) would ship with no wards at all: %s"
                         % ", ".join(sorted(empty)))
    for slug, d in out.items():
        for k in d:
            if k.startswith("_"):
                raise SystemExit("%s: ward key %r collides with the reserved "
                                 "metadata key" % (slug, k))
        d["_source"] = {
            "election": ELECTION,
            "electionDate": ELECTION_DATE,
            "provisional": PROVISIONAL,
            "asOf": str(datetime.date.today()),
            "publisher": PUBLISHER,
            "sourceUrl": SOURCE_URL,
        }
    return out, {"ltsb": len(ltsb), "paired": paired, "fallback": fallback,
                 "unpaired": unpaired, "notes": notes, "places": len(places),
                 "rows": len(body)}


def payload(d):
    return json.dumps(d, ensure_ascii=False, separators=(",", ":"), sort_keys=True) + "\n"


def report(stats, out, stream=sys.stderr):
    n_unpaired = sum(len(v) for v in stats["unpaired"].values())
    print("  WEC rows %d -> %d distinct polling places; LTSB wards %d, paired "
          "%d (%.2f%%), %d via the multi-county municipality fallback"
          % (stats["rows"], stats["places"], stats["ltsb"], stats["paired"],
             100.0 * stats["paired"] / stats["ltsb"], stats["fallback"]),
          file=stream)
    for n in stats["notes"]:
        print("  note: %s" % n, file=stream)
    print("  %d ward(s) in %d municipalities have no polling place in the "
          "Commission's file — the card degrades to MyVote for them:"
          % (n_unpaired, len(stats["unpaired"])), file=stream)
    for k, v in sorted(stats["unpaired"].items()):
        print("    %s / %s %s: wards %s"
              % (k[0].title(), k[1], k[2].title(),
                 ", ".join(sorted(v, key=ward_sort))), file=stream)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--check", action="store_true",
                    help="verify the shipped files, write nothing")
    ap.add_argument("--xlsx", default=DEFAULT_XLSX,
                    help="the Commission's workbook (default: the committed one)")
    args = ap.parse_args()

    out, stats = build(args.xlsx)

    def comparable(d):
        """Everything but the stamp the clock writes: `asOf` is the day the
        build ran, so --check would fail every day after the commit if it
        compared bytes. The PAIRING is what must not drift, and it does."""
        d = {k: v for k, v in d.items()}
        d["_source"] = {k: v for k, v in d["_source"].items() if k != "asOf"}
        return d

    total, stale = 0, []
    for slug in sorted(out):
        path = os.path.join(APP_DATA_DIR, "%s-polling-places.json" % slug)
        body = payload(out[slug])
        total += len(body)
        if args.check:
            if not os.path.exists(path):
                stale.append("%s-polling-places.json is missing" % slug)
                continue
            try:
                have = json.load(open(path, encoding="utf-8"))
            except Exception as e:
                stale.append("%s-polling-places.json does not parse: %s" % (slug, e))
                continue
            if "_source" not in have:
                stale.append("%s-polling-places.json carries no _source block" % slug)
            elif comparable(have) != comparable(out[slug]):
                stale.append("%s-polling-places.json differs from the "
                             "Commission's file" % slug)
        else:
            with open(path, "w", encoding="utf-8") as f:
                f.write(body)

    if args.check:
        if stale:
            for s in stale[:8]:
                print("  %s" % s, file=sys.stderr)
            raise SystemExit("build-wi-polling-places: FAIL — %d county file(s) "
                             "stale; regenerate with "
                             "wi/scripts/build_wi_polling_places.py" % len(stale))
        print("build-wi-polling-places: OK — %d counties, %d wards paired "
              "(%.2f%%), %.0f KB total, %s%s"
              % (len(out), stats["paired"], 100.0 * stats["paired"] / stats["ltsb"],
                 total / 1024.0, ELECTION,
                 " (PROVISIONAL)" if PROVISIONAL else ""))
        return

    print("build-wi-polling-places: wrote %d county files, %.0f KB total"
          % (len(out), total / 1024.0), file=sys.stderr)
    report(stats, out)


if __name__ == "__main__":
    main()
